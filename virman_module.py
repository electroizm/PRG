"""
Virman Modülü - Virman işlemlerini yönetir
"""

import os
import pyodbc
from datetime import datetime
import pandas as pd
import requests
from io import BytesIO
import pickle
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, QProgressBar,
                             QTableWidget, QTableWidgetItem, QHeaderView, QPushButton, QMessageBox, QShortcut)
from PyQt5.QtGui import QColor, QFont, QKeySequence

# Central config import
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from central_config import CentralConfigManager


class VirmanModule(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Virman Yönetimi")
        self.setGeometry(200, 200, 800, 600)

        # SQL bağlantı bilgileri - Environment variable'lardan al
        self.sql_connection = None
        self.sql_server = os.getenv('SQL_SERVER', '')
        self.sql_database = os.getenv('SQL_DATABASE', '')
        self.sql_username = os.getenv('SQL_USERNAME', '')
        self.sql_password = os.getenv('SQL_PASSWORD', '')

        # Seçili kasa ve ay bilgileri
        self.selected_kasa_kodu = None
        self.selected_kasa_adi = None
        self.selected_ay = None
        self.selected_yil = None
        self.selected_ay_adi = None

        # Lazy loading için flag
        self._data_loaded = False

        self.init_ui()

    def showEvent(self, event):
        """Widget ilk gösterildiğinde veri yükle (lazy loading)"""
        super().showEvent(event)
        if not self._data_loaded:
            self._data_loaded = True
            # UI render olduktan sonra veri yükle
            QTimer.singleShot(100, self.load_data_async)

    def load_data_async(self):
        """Verileri asenkron şekilde yükle"""
        self.status_label.setText("📊 Veriler yükleniyor...")
        from PyQt5.QtWidgets import QApplication
        QApplication.processEvents()  # UI'ın yanıt vermesini sağla
        self.get_kasalar()

    def init_ui(self):
        """UI'ı başlat"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        # Progress Bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 1px solid #2d2d2d;
                border-radius: 3px;
                background-color: #1a1a1a;
                color: #ffffff;
                text-align: center;
                font-weight: bold;
                min-height: 17px;
                max-height: 17px;
                font-size: 17px;
            }
            QProgressBar::chunk {
                background: qlineargradient(x1: 0, y1: 0, x2: 1, y2: 0, stop: 0 #4CAF50, stop: 1 #45a049);
                border-radius: 3px;
            }
        """)

        # Buton satırı (Sol üstte)
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)

        # Tümü Seç butonu (Gri)
        self.select_all_btn = QPushButton("Tümü Seç")
        self.select_all_btn.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border: none;
                padding: 8px 16px;
                font-weight: bold;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        self.select_all_btn.clicked.connect(self.select_all_rows)
        button_layout.addWidget(self.select_all_btn)

        # Mikro Güncelle butonu (Turuncu)
        self.mikro_btn = QPushButton("Mikro Güncelle")
        self.mikro_btn.setStyleSheet("""
            QPushButton {
                background-color: #ff8c00;
                color: white;
                border: none;
                padding: 8px 16px;
                font-weight: bold;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #ff7700;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
        """)
        self.mikro_btn.clicked.connect(self.run_mikro)
        button_layout.addWidget(self.mikro_btn)

        # Kaydet butonu (Kırmızı)
        self.save_btn = QPushButton("Kaydet")
        self.save_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                border: none;
                padding: 8px 16px;
                font-weight: bold;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
        """)
        self.save_btn.clicked.connect(self.save_virman_data)
        button_layout.addWidget(self.save_btn)

        # Yenile butonu (Yeşil)
        self.refresh_btn = QPushButton("Yenile")
        self.refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: none;
                padding: 8px 16px;
                font-weight: bold;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #218838;
            }
        """)
        self.refresh_btn.clicked.connect(self.refresh_data)
        button_layout.addWidget(self.refresh_btn)

        # Excel butonu (Mavi)
        self.excel_btn = QPushButton("Excel")
        self.excel_btn.setStyleSheet("""
            QPushButton {
                background-color: #007bff;
                color: white;
                border: none;
                padding: 8px 16px;
                font-weight: bold;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
        """)
        self.excel_btn.clicked.connect(self.export_detail_to_excel)
        button_layout.addWidget(self.excel_btn)

        # Boşluk ekle (butonlar solda kalsın)
        button_layout.addStretch()

        layout.addLayout(button_layout)

        # Dinamik ay isimleri
        ay_isimleri = [
            "OCAK", "ŞUBAT", "MART", "NİSAN", "MAYIS", "HAZİRAN",
            "TEMMUZ", "AĞUSTOS", "EYLÜL", "EKİM", "KASIM", "ARALIK"
        ]

        bugun = datetime.now()
        bu_ay = ay_isimleri[bugun.month - 1]  # Şimdiki ay

        # Önceki ayı hesapla
        if bugun.month == 1:
            onceki_ay = ay_isimleri[11]  # Aralık
        else:
            onceki_ay = ay_isimleri[bugun.month - 2]

        # Kasa Tablosu
        self.kasa_table = QTableWidget()
        self.kasa_table.setColumnCount(8)
        self.kasa_table.setHorizontalHeaderLabels([
            "Seç",
            "KASA KODU",
            "KASA ADI",
            "BAKİYE",
            "Virman",
            "Virman Bakiye",
            f"{onceki_ay} Bakiye",
            f"{bu_ay} Bakiye"
        ])

        # Tablo stilini ayarla - parent stylesheet'i override et
        # NOT: QTableWidget::item için HİÇBİR stil tanımlamıyoruz
        # Bu sayede setData(Qt.BackgroundRole) çalışabilir
        self.kasa_table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                gridline-color: #d0d0d0;
                border: none;
            }
            QHeaderView::section {
                background-color: #000000;
                color: white;
                padding: 10px;
                border: none;
                font-weight: bold;
                font-size: 13px;
            }
        """)

        # Sütun genişliklerini ayarla - tüm sütunlar veriye göre genişlesin
        header = self.kasa_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeToContents)
        header.setStretchLastSection(False)

        # Satır yüksekliğini ayarla
        self.kasa_table.verticalHeader().setDefaultSectionSize(35)
        self.kasa_table.verticalHeader().setVisible(False)

        # Scroll bar'ı kaldır - tablo dinamik yükseklikte olacak
        self.kasa_table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.kasa_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)

        # Tablo item click event'i bağla
        self.kasa_table.itemClicked.connect(self.on_table_item_clicked)

        # Ctrl+C kısayolu - Kasa tablosu
        self.copy_shortcut_kasa = QShortcut(QKeySequence("Ctrl+C"), self.kasa_table)
        self.copy_shortcut_kasa.activated.connect(lambda: self.copy_table_selection(self.kasa_table))

        layout.addWidget(self.kasa_table)  # Ana tablo - stretch faktörü yok (dinamik yükseklik)

        # Detay bölümü için horizontal layout (Giriş ve Çıkış tabloları)
        self.detail_layout = QHBoxLayout()
        self.detail_layout.setSpacing(10)

        # Giriş Tablosu (Sol)
        self.giris_table = QTableWidget()
        self.giris_table.setColumnCount(5)
        self.giris_table.setHorizontalHeaderLabels(["Tarih", "KASA ADI", "CARI ADI", "Tutar", "Açıklama"])
        self.giris_table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                gridline-color: #d0d0d0;
                border: 1px solid #28a745;
            }
            QHeaderView::section {
                background-color: #28a745;
                color: white;
                padding: 8px;
                border: none;
                font-weight: bold;
                font-size: 12px;
            }
        """)
        # Sütun genişlik ayarları: Tarih, KASA ADI, CARI ADI, Tutar -> dinamik, Açıklama -> kalan alan
        giris_header = self.giris_table.horizontalHeader()
        giris_header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # Tarih
        giris_header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # KASA ADI
        giris_header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # CARI ADI
        giris_header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # Tutar
        giris_header.setSectionResizeMode(4, QHeaderView.Stretch)  # Açıklama - kalan alan
        self.giris_table.verticalHeader().setVisible(False)
        self.giris_table.setMinimumHeight(250)  # Minimum yükseklik
        self.giris_table.setVisible(False)  # Başlangıçta gizli
        
        # Ctrl+C kısayolu - Giriş tablosu
        self.copy_shortcut_giris = QShortcut(QKeySequence("Ctrl+C"), self.giris_table)
        self.copy_shortcut_giris.activated.connect(lambda: self.copy_table_selection(self.giris_table))
        
        self.detail_layout.addWidget(self.giris_table)

        # Çıkış Tablosu (Sağ)
        self.cikis_table = QTableWidget()
        self.cikis_table.setColumnCount(5)
        self.cikis_table.setHorizontalHeaderLabels(["Tarih", "KASA ADI", "CARI ADI", "Tutar", "Açıklama"])
        self.cikis_table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                gridline-color: #d0d0d0;
                border: 1px solid #dc3545;
            }
            QHeaderView::section {
                background-color: #dc3545;
                color: white;
                padding: 8px;
                border: none;
                font-weight: bold;
                font-size: 12px;
            }
        """)
        # Sütun genişlik ayarları: Tarih, KASA ADI, CARI ADI, Tutar -> dinamik, Açıklama -> kalan alan
        cikis_header = self.cikis_table.horizontalHeader()
        cikis_header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # Tarih
        cikis_header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # KASA ADI
        cikis_header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # CARI ADI
        cikis_header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # Tutar
        cikis_header.setSectionResizeMode(4, QHeaderView.Stretch)  # Açıklama - kalan alan
        self.cikis_table.verticalHeader().setVisible(False)
        self.cikis_table.setMinimumHeight(250)  # Minimum yükseklik
        self.cikis_table.setVisible(False)  # Başlangıçta gizli
        
        # Ctrl+C kısayolu - Çıkış tablosu
        self.copy_shortcut_cikis = QShortcut(QKeySequence("Ctrl+C"), self.cikis_table)
        self.copy_shortcut_cikis.activated.connect(lambda: self.copy_table_selection(self.cikis_table))
        
        self.detail_layout.addWidget(self.cikis_table)

        layout.addLayout(self.detail_layout, 1)  # Detay tabloları - kalan alanı alsın

        # Status Layout
        status_layout = QHBoxLayout()

        self.status_label = QLabel("Hazır")
        self.status_label.setStyleSheet("""
            QLabel {
                color: #cccccc;
                padding: 4px 8px;
                background-color: #2d2d2d;
                border-top: 1px solid #404040;
                font-size: 14px;
                max-height: 20px;
            }
        """)

        status_layout.addWidget(self.status_label, 3)
        status_layout.addWidget(self.progress_bar, 1)
        status_layout.setContentsMargins(0, 0, 0, 0)

        status_widget = QWidget()
        status_widget.setLayout(status_layout)
        status_widget.setStyleSheet("background-color: #2d2d2d; border-top: 1px solid #404040;")

        layout.addWidget(status_widget)

        # NOT: Widget-level stylesheet REMOVED to prevent cascading to table items

    def format_tutar(self, tutar_str):
        """Tutarı finansal formata çevir (örn: 1.740.676 ₺)"""
        try:
            # String'i float'a çevir
            tutar = float(tutar_str) if tutar_str else 0.0
            # Int'e yuvarla
            tutar_int = int(round(tutar))
            # Binlik ayraçlarla formatla
            formatted = f"{tutar_int:,}".replace(",", ".")
            return f"{formatted} ₺"
        except (ValueError, TypeError):
            return "0 ₺"

    def parse_tutar(self, tutar_str):
        """Tutar string'ini float'a çevir: '1.740.676 ₺' -> 1740676.0"""
        try:
            if not tutar_str or tutar_str.strip() == "":
                return 0.0
            # ₺ ve boşlukları temizle
            clean = str(tutar_str).replace(" ₺", "").replace("₺", "").strip()
            # Binlik ayraçları temizle (.)
            clean = clean.replace(".", "")
            # Virgül varsa nokta yap (ondalık için)
            clean = clean.replace(",", ".")
            return float(clean) if clean else 0.0
        except (ValueError, TypeError):
            return 0.0

    def get_virman_data(self):
        """PRGsheet/Virman sayfasından virman verilerini çek"""
        try:
            config_manager = CentralConfigManager()
            spreadsheet_id = config_manager.MASTER_SPREADSHEET_ID

            # Google Sheets'i Excel formatında indir
            url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
            response = requests.get(url, timeout=10)

            if response.status_code != 200:
                self.status_label.setText("Virman sayfası okunamadı")
                return {}

            # Virman sayfasını oku
            df = pd.read_excel(BytesIO(response.content), sheet_name="Virman")

            # KASA KODU -> Virman değeri mapping'i oluştur
            virman_dict = {}
            for _, row in df.iterrows():
                kasa_kodu = str(row.get('KASA KODU', '')).strip()
                virman_value = row.get('Virman', 0)
                if kasa_kodu:
                    virman_dict[kasa_kodu] = virman_value

            return virman_dict

        except Exception as e:
            self.status_label.setText(f"Virman verisi çekme hatası: {str(e)}")
            return {}

    def get_kasa_monthly_data(self):
        """PRGsheet/Kasa sayfasından aylık bakiye verilerini çek"""
        try:
            config_manager = CentralConfigManager()
            spreadsheet_id = config_manager.MASTER_SPREADSHEET_ID

            # Google Sheets'i Excel formatında indir
            url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
            response = requests.get(url, timeout=10)

            if response.status_code != 200:
                self.status_label.setText("Kasa sayfası okunamadı")
                return {}

            # Kasa sayfasını oku
            df = pd.read_excel(BytesIO(response.content), sheet_name="Kasa")

            # Tarih sütununu datetime'a çevir (eğer varsa)
            if 'Tarih' in df.columns:
                df['Tarih'] = pd.to_datetime(df['Tarih'], errors='coerce')

            # KASA KODU ve Ay bazında gruplama yapmak için dictionary oluştur
            # Yapı: {KASA_KODU: {ay_no: toplam_tutar}}
            kasa_monthly_dict = {}

            for _, row in df.iterrows():
                kasa_kodu = str(row.get('KASA KODU', '')).strip()
                tarih = row.get('Tarih')
                tutar = row.get('TUTAR', 0)

                if pd.isna(tarih) or not kasa_kodu:
                    continue

                # Ay ve yıl bilgisini al
                ay = tarih.month
                yil = tarih.year

                # Dictionary'de kasa kodu yoksa ekle
                if kasa_kodu not in kasa_monthly_dict:
                    kasa_monthly_dict[kasa_kodu] = {}

                # Ay bazında topla
                ay_key = f"{yil}-{ay:02d}"
                if ay_key not in kasa_monthly_dict[kasa_kodu]:
                    kasa_monthly_dict[kasa_kodu][ay_key] = 0

                try:
                    kasa_monthly_dict[kasa_kodu][ay_key] += float(tutar) if tutar else 0
                except (ValueError, TypeError):
                    pass

            return kasa_monthly_dict

        except Exception as e:
            self.status_label.setText(f"Kasa aylık verisi çekme hatası: {str(e)}")
            return {}

    def connect_to_sql(self):
        """SQL Server'a bağlan"""
        try:
            connection_string = (
                f"DRIVER={{SQL Server}};"
                f"SERVER={self.sql_server};"
                f"DATABASE={self.sql_database};"
                f"UID={self.sql_username};"
                f"PWD={self.sql_password};"
            )

            self.sql_connection = pyodbc.connect(connection_string)
            self.status_label.setText("SQL Server'a bağlandı")

            # Kasaları otomatik çek
            self.get_kasalar()
            return True

        except Exception as e:
            self.status_label.setText(f"SQL bağlantı hatası: {str(e)}")
            return False

    def disconnect_sql(self):
        """SQL bağlantısını kapat"""
        try:
            if self.sql_connection:
                self.sql_connection.close()
                self.sql_connection = None
                self.status_label.setText("Bağlantı kapatıldı")
        except Exception as e:
            self.status_label.setText(f"Bağlantı kapatma hatası: {str(e)}")

    def get_kasalar(self):
        """Nakit kasalarını PRGsheet/Bakiye sayfasından çek"""
        try:
            from PyQt5.QtWidgets import QApplication

            # PRGsheet/Virman sayfasından virman verilerini çek
            self.status_label.setText("📊 Virman verileri yükleniyor...")
            QApplication.processEvents()
            virman_data = self.get_virman_data()

            # PRGsheet/Kasa sayfasından aylık bakiye verilerini çek
            self.status_label.setText("📊 Aylık bakiye verileri yükleniyor...")
            QApplication.processEvents()
            kasa_monthly_data = self.get_kasa_monthly_data()

            # Dinamik ay bilgilerini hesapla
            bugun = datetime.now()
            bu_ay_no = bugun.month
            bu_yil = bugun.year

            # Önceki ayı hesapla
            if bu_ay_no == 1:
                onceki_ay_no = 12
                onceki_yil = bu_yil - 1
            else:
                onceki_ay_no = bu_ay_no - 1
                onceki_yil = bu_yil

            # Ay key'leri oluştur (YYYY-MM formatında)
            onceki_ay_key = f"{onceki_yil}-{onceki_ay_no:02d}"
            bu_ay_key = f"{bu_yil}-{bu_ay_no:02d}"

            # PRGsheet/Bakiye sayfasından veri çek
            self.status_label.setText("📊 Bakiye verileri yükleniyor...")
            QApplication.processEvents()
            config_manager = CentralConfigManager()
            spreadsheet_id = config_manager.MASTER_SPREADSHEET_ID

            url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
            response = requests.get(url, timeout=10)

            if response.status_code != 200:
                self.status_label.setText("Bakiye sayfası okunamadı")
                return None

            # Bakiye sayfasını oku
            df = pd.read_excel(BytesIO(response.content), sheet_name="Bakiye")

            # Tabloyu temizle
            self.kasa_table.setRowCount(0)

            # Tabloya verileri ekle
            self.status_label.setText("📊 Tablo dolduruluyor...")
            QApplication.processEvents()
            kasalar = []
            table_row_idx = 0  # Tablo satır indeksi
            for _, row in df.iterrows():
                # Sütun adlarını kullanarak verileri al
                kasa_kodu = str(row.get('KASA KODU', '')).strip()
                kasa_ismi = str(row.get('KASA ADI', '')).strip()
                bakiye_raw = row.get('BAKIYE', 0)

                # Boş satırları atla
                if not kasa_kodu:
                    continue

                # Tablo satırı ekle
                self.kasa_table.insertRow(table_row_idx)

                # Virman değerini PRGsheet'ten al
                virman_value = virman_data.get(kasa_kodu, 0)

                # Aylık bakiye verilerini PRGsheet/Kasa'dan al
                onceki_ay_bakiye = 0
                bu_ay_bakiye = 0
                if kasa_kodu in kasa_monthly_data:
                    onceki_ay_bakiye = kasa_monthly_data[kasa_kodu].get(onceki_ay_key, 0)
                    bu_ay_bakiye = kasa_monthly_data[kasa_kodu].get(bu_ay_key, 0)

                # Bold font oluştur
                bold_font = QFont()
                bold_font.setBold(True)

                # Checkbox (Seç sütunu)
                checkbox_item = QTableWidgetItem()
                checkbox_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
                checkbox_item.setCheckState(Qt.Unchecked)
                self.kasa_table.setItem(table_row_idx, 0, checkbox_item)

                # KASA KODU (Düzenlenemez)
                item_kodu = QTableWidgetItem(kasa_kodu)
                item_kodu.setFont(bold_font)
                item_kodu.setForeground(QColor(0, 0, 0))
                item_kodu.setFlags(item_kodu.flags() & ~Qt.ItemIsEditable)  # Düzenlenemez
                self.kasa_table.setItem(table_row_idx, 1, item_kodu)

                # KASA ADI (Düzenlenemez)
                item_ismi = QTableWidgetItem(kasa_ismi)
                item_ismi.setFont(bold_font)
                item_ismi.setForeground(QColor(0, 0, 0))
                item_ismi.setFlags(item_ismi.flags() & ~Qt.ItemIsEditable)  # Düzenlenemez
                self.kasa_table.setItem(table_row_idx, 2, item_ismi)

                # Virman Bakiye hesapla: Virman + Bu Ay (EKİM Bakiye hesaplamaya dahil değil)
                virman_bakiye_toplam = virman_value + bu_ay_bakiye

                # BAKİYE - Formatlanmış (Düzenlenemez)
                bakiye_formatli = self.format_tutar(bakiye_raw)
                item_bakiye = QTableWidgetItem(bakiye_formatli)
                item_bakiye.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                item_bakiye.setFont(bold_font)
                item_bakiye.setFlags(item_bakiye.flags() & ~Qt.ItemIsEditable)  # Düzenlenemez

                # Fark hesapla ve renklendirme
                bakiye_float = float(bakiye_raw) if bakiye_raw else 0.0
                fark = abs(bakiye_float - virman_bakiye_toplam)

                # Renklendirme - setData ile explicit role kullan
                if fark < 70:
                    item_bakiye.setData(Qt.BackgroundRole, QColor(144, 238, 144))  # Yeşil
                else:
                    item_bakiye.setData(Qt.BackgroundRole, QColor(255, 182, 193))  # Kırmızı

                item_bakiye.setData(Qt.ForegroundRole, QColor(0, 0, 0))  # Siyah text

                self.kasa_table.setItem(table_row_idx, 3, item_bakiye)

                # Virman (PRGsheet/Virman sayfasından) - DÜZENLENEBİLİR
                virman_formatli = self.format_tutar(virman_value)
                item_virman = QTableWidgetItem(virman_formatli)
                item_virman.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                item_virman.setFont(bold_font)
                item_virman.setForeground(QColor(0, 0, 0))
                # Virman sütunu düzenlenebilir - flag değiştirmiyoruz
                self.kasa_table.setItem(table_row_idx, 4, item_virman)

                # Virman Bakiye (Düzenlenemez)
                virman_bakiye_formatli = self.format_tutar(virman_bakiye_toplam)
                item_virman_bakiye = QTableWidgetItem(virman_bakiye_formatli)
                item_virman_bakiye.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                item_virman_bakiye.setFont(bold_font)
                item_virman_bakiye.setFlags(item_virman_bakiye.flags() & ~Qt.ItemIsEditable)  # Düzenlenemez

                # Virman Bakiye için aynı renklendirme - setData ile explicit role kullan
                if fark < 70:
                    item_virman_bakiye.setData(Qt.BackgroundRole, QColor(144, 238, 144))
                else:
                    item_virman_bakiye.setData(Qt.BackgroundRole, QColor(255, 182, 193))

                item_virman_bakiye.setData(Qt.ForegroundRole, QColor(0, 0, 0))

                self.kasa_table.setItem(table_row_idx, 5, item_virman_bakiye)

                # Önceki Ay Bakiye (Düzenlenemez)
                onceki_ay_formatli = self.format_tutar(onceki_ay_bakiye)
                item_onceki_ay = QTableWidgetItem(onceki_ay_formatli)
                item_onceki_ay.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                item_onceki_ay.setFont(bold_font)
                item_onceki_ay.setForeground(QColor(0, 0, 0))
                item_onceki_ay.setFlags(item_onceki_ay.flags() & ~Qt.ItemIsEditable)  # Düzenlenemez
                self.kasa_table.setItem(table_row_idx, 6, item_onceki_ay)

                # Bu Ay Bakiye (Düzenlenemez)
                bu_ay_formatli = self.format_tutar(bu_ay_bakiye)
                item_bu_ay = QTableWidgetItem(bu_ay_formatli)
                item_bu_ay.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                item_bu_ay.setFont(bold_font)
                item_bu_ay.setForeground(QColor(0, 0, 0))
                item_bu_ay.setFlags(item_bu_ay.flags() & ~Qt.ItemIsEditable)  # Düzenlenemez
                self.kasa_table.setItem(table_row_idx, 7, item_bu_ay)

                # Kasalar listesine ekle
                row_dict = {
                    'KASA KODU': kasa_kodu,
                    'KASA ADI': kasa_ismi,
                    'BAKIYE': bakiye_raw
                }
                kasalar.append(row_dict)

                # Tablo satır indeksini artır
                table_row_idx += 1

            # Tablonun yüksekliğini satır sayısına göre ayarla (scroll bar olmaması için)
            row_count = self.kasa_table.rowCount()
            row_height = 35  # Satır yüksekliği
            header_height = self.kasa_table.horizontalHeader().height()
            total_height = header_height + (row_count * row_height) + 5  # +5 padding
            self.kasa_table.setMinimumHeight(total_height)
            self.kasa_table.setMaximumHeight(total_height)

            self.status_label.setText(f"✅ {len(df)} kasa başarıyla yüklendi")

            return kasalar

        except Exception as e:
            self.status_label.setText(f"Kasa çekme hatası: {str(e)}")
            return None

    def select_all_rows(self):
        """Tüm satırları seç/seçimi kaldır"""
        # İlk satırın checkbox durumuna göre toggle yap
        if self.kasa_table.rowCount() == 0:
            return

        first_checkbox = self.kasa_table.item(0, 0)
        if first_checkbox is None:
            return

        # Eğer ilk checkbox işaretliyse, tümünün işaretini kaldır
        # Değilse, tümünü işaretle
        new_state = Qt.Unchecked if first_checkbox.checkState() == Qt.Checked else Qt.Checked

        for row in range(self.kasa_table.rowCount()):
            checkbox_item = self.kasa_table.item(row, 0)
            if checkbox_item:
                checkbox_item.setCheckState(new_state)

        state_text = "işaretlendi" if new_state == Qt.Checked else "işaret kaldırıldı"
        self.status_label.setText(f"Tüm satırlar {state_text}")

    def save_virman_data(self):
        """Tablodaki seçili satırların Virman sütunu değerlerini PRGsheet/Virman sayfasına kaydet"""
        try:
            # Service Account ile Google Sheets client'ı al
            config_manager = CentralConfigManager()
            gc = config_manager.gc

            # PRGsheet'i aç
            spreadsheet = gc.open("PRGsheet")
            virman_worksheet = spreadsheet.worksheet('Virman')

            # Önce mevcut Virman sayfasını oku (A: KASA KODU, B: KASA ADI, C: Virman)
            values = virman_worksheet.get_all_values()

            if not values:
                self.status_label.setText("Virman sayfası boş!")
                return

            # Header'ı atla ve mevcut verileri dictionary'e çevir
            virman_row_map = {}  # {KASA_KODU: row_index}

            for idx, row in enumerate(values[1:], start=2):  # 2'den başla (header + 0-index)
                if len(row) > 0:
                    kasa_kodu = str(row[0]).strip()
                    virman_row_map[kasa_kodu] = idx

            # Tablodaki SEÇİLİ satırların Virman sütununu oku ve güncelleme listesi oluştur
            updates = []
            selected_count = 0

            for row_idx in range(self.kasa_table.rowCount()):
                # Checkbox kontrolü - sadece seçili satırları işle
                checkbox_item = self.kasa_table.item(row_idx, 0)
                if not checkbox_item or checkbox_item.checkState() != Qt.Checked:
                    continue  # Seçili değilse atla

                selected_count += 1

                kasa_kodu_item = self.kasa_table.item(row_idx, 1)  # KASA KODU
                virman_item = self.kasa_table.item(row_idx, 4)  # Virman sütunu

                if kasa_kodu_item and virman_item:
                    kasa_kodu = kasa_kodu_item.text()
                    virman_text = virman_item.text()

                    # "1.740.676 ₺" formatından sayıya çevir
                    try:
                        # ₺ sembolünü ve boşlukları kaldır
                        virman_clean = virman_text.replace(' ₺', '').replace('₺', '').strip()
                        # Nokta ayraçlarını kaldır
                        virman_clean = virman_clean.replace('.', '')
                        # Virgülü noktaya çevir (varsa)
                        virman_clean = virman_clean.replace(',', '.')
                        virman_value = float(virman_clean) if virman_clean else 0
                    except:
                        virman_value = 0

                    # Bu KASA KODU için Virman sayfasındaki satırı bul
                    if kasa_kodu in virman_row_map:
                        row_number = virman_row_map[kasa_kodu]
                        # C sütununa yaz (Virman sütunu - 3. sütun)
                        updates.append({
                            'range': f'Virman!C{row_number}',
                            'values': [[virman_value]]
                        })

            if selected_count == 0:
                self.status_label.setText("Hiçbir satır seçilmedi!")
                QMessageBox.warning(self, "Uyarı", "Lütfen güncellemek istediğiniz satırları seçin!")
                return

            if not updates:
                self.status_label.setText("Güncellenecek veri bulunamadı")
                return

            # Batch update yap - gspread ile
            for update in updates:
                cell_range = update['range'].replace('Virman!', '')  # 'C2' formatına çevir
                value = update['values'][0][0]
                virman_worksheet.update(cell_range, [[value]], value_input_option='RAW')
        
        # Progress güncelleme
            updated_cells = len(updates)
            self.status_label.setText(f"✅ {updated_cells} hücre güncellendi ({selected_count} satır)")
            QMessageBox.information(self, "Başarılı", f"{updated_cells} hücre başarıyla güncellendi!\n({selected_count} satır)")

        except Exception as e:
            self.status_label.setText(f"Kaydetme hatası: {str(e)}")
            QMessageBox.critical(self, "Hata", f"Virman verileri kaydedilemedi:\n{str(e)}")

    def copy_table_selection(self, table):
        """Tablodaki seçili hücreyi/hücreleri kopyala"""
        from PyQt5.QtWidgets import QApplication
        
        selected_items = table.selectedItems()
        if not selected_items:
            return

        # Sadece ilk seçili öğeyi kopyala (basitlik için)
        text = selected_items[0].text()
        if text:
            QApplication.clipboard().setText(text)
            old_text = self.status_label.text()
            self.status_label.setText("✅ Kopyalandı")
            QTimer.singleShot(1500, lambda t=old_text: self.status_label.setText(t))
        else:
            self.status_label.setText("⚠️ Boş hücre")

    def refresh_data(self):
        """Tabloyu yenile"""
        self.status_label.setText("Veriler yenileniyor...")
        self.get_kasalar()
        self.status_label.setText("Veriler yenilendi")

    def on_table_item_clicked(self, item):
        """Tablo hücresine tıklandığında çağrılır"""
        row = item.row()
        col = item.column()

        # Sadece 6. (EKİM Bakiye) veya 7. (KASIM Bakiye) sütunlarına tıklanırsa işlem yap
        if col not in [6, 7]:
            return

        # KASA KODU'nu al
        kasa_kodu_item = self.kasa_table.item(row, 1)
        if not kasa_kodu_item:
            return

        kasa_kodu = kasa_kodu_item.text()

        # KASA ADI'nı al
        kasa_adi_item = self.kasa_table.item(row, 2)
        kasa_adi = kasa_adi_item.text() if kasa_adi_item else kasa_kodu

        # Hangi aya tıklandığını belirle
        bugun = datetime.now()
        bu_yil = bugun.year

        if col == 6:  # Önceki ay
            if bugun.month == 1:
                ay_no = 12
                yil = bu_yil - 1
            else:
                ay_no = bugun.month - 1
                yil = bu_yil
        else:  # col == 7, Bu ay
            ay_no = bugun.month
            yil = bu_yil

        # Ay adını al
        ay_isimleri = [
            "OCAK", "ŞUBAT", "MART", "NİSAN", "MAYIS", "HAZİRAN",
            "TEMMUZ", "AĞUSTOS", "EYLÜL", "EKİM", "KASIM", "ARALIK"
        ]
        ay_adi = ay_isimleri[ay_no - 1]

        # Seçili bilgileri kaydet (Excel export için)
        self.selected_kasa_kodu = kasa_kodu
        self.selected_kasa_adi = kasa_adi
        self.selected_ay = ay_no
        self.selected_yil = yil
        self.selected_ay_adi = ay_adi

        # Detayları yükle
        self.load_kasa_details(kasa_kodu, yil, ay_no, ay_adi)

    def load_kasa_details(self, kasa_kodu, yil, ay_no, ay_adi):
        """Belirli bir kasa ve ay için detayları PRGsheet/Kasa'dan çek ve göster"""
        try:
            config_manager = CentralConfigManager()
            spreadsheet_id = config_manager.MASTER_SPREADSHEET_ID

            # Google Sheets'i Excel formatında indir
            url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
            response = requests.get(url, timeout=10)

            if response.status_code != 200:
                self.status_label.setText("Kasa detayları okunamadı")
                return

            # Kasa sayfasını oku
            df = pd.read_excel(BytesIO(response.content), sheet_name="Kasa")

            # Tarih sütununu datetime'a çevir
            if 'Tarih' in df.columns:
                df['Tarih'] = pd.to_datetime(df['Tarih'], errors='coerce')

            # Filtreleme: KASA KODU ve Ay
            filtered_df = df[
                (df['KASA KODU'].astype(str).str.strip() == kasa_kodu) &
                (df['Tarih'].dt.year == yil) &
                (df['Tarih'].dt.month == ay_no)
            ]

            if filtered_df.empty:
                self.status_label.setText(f"{kasa_kodu} - {ay_adi} {yil} için veri bulunamadı")
                self.giris_table.setVisible(False)
                self.cikis_table.setVisible(False)
                return

            # Giriş ve Çıkışları ayır
            giris_df = filtered_df[filtered_df['TUTAR'] > 0].copy()
            cikis_df = filtered_df[filtered_df['TUTAR'] < 0].copy()

            # Giriş tablosunu doldur
            self.giris_table.setRowCount(len(giris_df))
            for table_row, (_, row) in enumerate(giris_df.iterrows()):
                tarih = row['Tarih'].strftime('%d.%m.%Y') if pd.notna(row['Tarih']) else ''
                kasa_adi = str(row.get('KASA ADI', '')) if pd.notna(row.get('KASA ADI')) else ''
                cari_adi = str(row.get('CARI ADI', '')) if pd.notna(row.get('CARI ADI')) else ''
                aciklama = str(row.get('ACIKLAMA', '')) if pd.notna(row.get('ACIKLAMA')) else ''
                tutar = self.format_tutar(row['TUTAR'])

                bold_font = QFont()
                bold_font.setBold(True)

                # Tarih (Sütun 0)
                tarih_item = QTableWidgetItem(tarih)
                tarih_item.setFont(bold_font)
                tarih_item.setForeground(QColor(0, 0, 0))  # Siyah text
                self.giris_table.setItem(table_row, 0, tarih_item)

                # KASA ADI (Sütun 1)
                kasa_adi_item = QTableWidgetItem(kasa_adi)
                kasa_adi_item.setFont(bold_font)
                kasa_adi_item.setForeground(QColor(0, 0, 0))  # Siyah text
                self.giris_table.setItem(table_row, 1, kasa_adi_item)

                # CARI ADI (Sütun 2)
                cari_adi_item = QTableWidgetItem(cari_adi)
                cari_adi_item.setFont(bold_font)
                cari_adi_item.setForeground(QColor(0, 0, 0))  # Siyah text
                self.giris_table.setItem(table_row, 2, cari_adi_item)

                # Tutar (Sütun 3)
                tutar_item = QTableWidgetItem(tutar)
                tutar_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                tutar_item.setFont(bold_font)
                tutar_item.setForeground(QColor(0, 0, 0))  # Siyah text
                self.giris_table.setItem(table_row, 3, tutar_item)

                # Açıklama (Sütun 4)
                aciklama_item = QTableWidgetItem(aciklama)
                aciklama_item.setFont(bold_font)
                aciklama_item.setForeground(QColor(0, 0, 0))  # Siyah text
                self.giris_table.setItem(table_row, 4, aciklama_item)

            # Çıkış tablosunu doldur
            self.cikis_table.setRowCount(len(cikis_df))
            for table_row, (_, row) in enumerate(cikis_df.iterrows()):
                tarih = row['Tarih'].strftime('%d.%m.%Y') if pd.notna(row['Tarih']) else ''
                kasa_adi = str(row.get('KASA ADI', '')) if pd.notna(row.get('KASA ADI')) else ''
                cari_adi = str(row.get('CARI ADI', '')) if pd.notna(row.get('CARI ADI')) else ''
                aciklama = str(row.get('ACIKLAMA', '')) if pd.notna(row.get('ACIKLAMA')) else ''
                tutar = self.format_tutar(abs(row['TUTAR']))  # Mutlak değer al

                bold_font = QFont()
                bold_font.setBold(True)

                # Tarih (Sütun 0)
                tarih_item = QTableWidgetItem(tarih)
                tarih_item.setFont(bold_font)
                tarih_item.setForeground(QColor(0, 0, 0))  # Siyah text
                self.cikis_table.setItem(table_row, 0, tarih_item)

                # KASA ADI (Sütun 1)
                kasa_adi_item = QTableWidgetItem(kasa_adi)
                kasa_adi_item.setFont(bold_font)
                kasa_adi_item.setForeground(QColor(0, 0, 0))  # Siyah text
                self.cikis_table.setItem(table_row, 1, kasa_adi_item)

                # CARI ADI (Sütun 2)
                cari_adi_item = QTableWidgetItem(cari_adi)
                cari_adi_item.setFont(bold_font)
                cari_adi_item.setForeground(QColor(0, 0, 0))  # Siyah text
                self.cikis_table.setItem(table_row, 2, cari_adi_item)

                # Tutar (Sütun 3)
                tutar_item = QTableWidgetItem(tutar)
                tutar_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                tutar_item.setFont(bold_font)
                tutar_item.setForeground(QColor(0, 0, 0))  # Siyah text
                self.cikis_table.setItem(table_row, 3, tutar_item)

                # Açıklama (Sütun 4)
                aciklama_item = QTableWidgetItem(aciklama)
                aciklama_item.setFont(bold_font)
                aciklama_item.setForeground(QColor(0, 0, 0))  # Siyah text
                self.cikis_table.setItem(table_row, 4, aciklama_item)

            # Tabloları görünür yap
            self.giris_table.setVisible(True)
            self.cikis_table.setVisible(True)

            giris_toplam = giris_df['TUTAR'].sum()
            cikis_toplam = abs(cikis_df['TUTAR'].sum())
            self.status_label.setText(
                f"{kasa_kodu} - {ay_adi} {yil} | "
                f"Giriş: {len(giris_df)} adet ({self.format_tutar(giris_toplam)}) | "
                f"Çıkış: {len(cikis_df)} adet ({self.format_tutar(cikis_toplam)})"
            )

        except Exception as e:
            self.status_label.setText(f"Detay yükleme hatası: {str(e)}")
            self.giris_table.setVisible(False)
            self.cikis_table.setVisible(False)

    def run_mikro(self):
        """Kasa.exe dosyasını çalıştır ve ardından verileri yenile"""
        try:
            exe_path = r"D:/GoogleDrive/PRG/EXE/Kasa.exe"
            if not os.path.exists(exe_path):
                self.status_label.setText(f"❌ Kasa.exe bulunamadı: {exe_path}")
                return

            self.status_label.setText("🔄 Kasa.exe çalıştırılıyor...")
            self.mikro_btn.setEnabled(False)

            os.startfile(exe_path)

            # 7 saniye sonra program bitmiş sayıp kontrol et
            QTimer.singleShot(7000, self.on_mikro_finished)

        except Exception as e:
            self.status_label.setText(f"❌ Program çalıştırma hatası: {str(e)}")
            self.mikro_btn.setEnabled(True)

    def on_mikro_finished(self):
        """Kasa.exe bittikten sonra Bakiye.exe'yi çalıştır"""
        try:
            exe_path = r"D:/GoogleDrive/PRG/EXE/Bakiye.exe"
            if not os.path.exists(exe_path):
                self.status_label.setText(f"❌ Bakiye.exe bulunamadı: {exe_path}")
                self.mikro_btn.setEnabled(True)
                return

            self.status_label.setText("✅ Kasa.exe tamamlandı, Bakiye.exe çalıştırılıyor...")
            os.startfile(exe_path)

            # 5 saniye sonra Bakiye.exe bitmiş sayıp devam et
            QTimer.singleShot(5000, self.on_bakiye_finished)

        except Exception as e:
            self.status_label.setText(f"❌ Bakiye.exe çalıştırma hatası: {str(e)}")
            self.mikro_btn.setEnabled(True)

    def on_bakiye_finished(self):
        """Bakiye.exe bittikten sonra 5 saniye bekle ve yenile"""
        self.status_label.setText("✅ Bakiye.exe tamamlandı, Google Sheets güncelleme bekleniyor...")

        # Google Sheets'e kaydedilmesi için ek bekleme (5 saniye)
        QTimer.singleShot(5000, self.delayed_data_refresh)

    def delayed_data_refresh(self):
        """Gecikmeli veri yenileme - kasa_module.py'deki Verileri Yenile butonu gibi"""
        self.status_label.setText("🔄 Veriler yenileniyor...")
        self.refresh_data()
        self.mikro_btn.setEnabled(True)

    def export_detail_to_excel(self):
        """Giriş ve Çıkış tablolarını yan yana aynı Excel sayfasına kaydet"""
        if not self.selected_kasa_adi or not self.selected_ay_adi:
            self.status_label.setText("⚠️ Lütfen önce bir kasa ve ay seçin (EKİM veya KASIM Bakiye'ye tıklayın)")
            QMessageBox.warning(self, "Uyarı", "Lütfen önce bir kasa ve ay seçin (EKİM veya KASIM Bakiye'ye tıklayın)")
            return

        if not self.giris_table.isVisible() and not self.cikis_table.isVisible():
            self.status_label.setText("⚠️ Dışa aktarılacak detay verisi yok")
            QMessageBox.warning(self, "Uyarı", "Dışa aktarılacak detay verisi yok")
            return

        try:
            # Dosya adı: {KASA_ADI}_{AY_ADI}_{YIL}
            file_name = f"{self.selected_kasa_adi}_{self.selected_ay_adi}_{self.selected_yil}"
            output_path = f"D:/GoogleDrive/~ {file_name}.xlsx"

            # Giriş tablosunu DataFrame'e çevir (Tutar'ı float olarak parse et)
            giris_data = []
            for row in range(self.giris_table.rowCount()):
                row_data = []
                for col in range(self.giris_table.columnCount()):
                    item = self.giris_table.item(row, col)
                    if col == 3:  # Tutar sütunu
                        row_data.append(self.parse_tutar(item.text() if item else "0"))
                    else:
                        row_data.append(item.text() if item else "")
                giris_data.append(row_data)

            giris_df = pd.DataFrame(giris_data, columns=["Tarih", "KASA ADI", "CARI ADI", "Tutar", "Açıklama"])

            # Çıkış tablosunu DataFrame'e çevir (Tutar'ı float olarak parse et)
            cikis_data = []
            for row in range(self.cikis_table.rowCount()):
                row_data = []
                for col in range(self.cikis_table.columnCount()):
                    item = self.cikis_table.item(row, col)
                    if col == 3:  # Tutar sütunu
                        row_data.append(self.parse_tutar(item.text() if item else "0"))
                    else:
                        row_data.append(item.text() if item else "")
                cikis_data.append(row_data)

            cikis_df = pd.DataFrame(cikis_data, columns=["Tarih", "KASA ADI", "CARI ADI", "Tutar", "Açıklama"])

            # Excel'e yan yana kaydet
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = f"{self.selected_kasa_kodu}_{self.selected_ay_adi}"

            # Başlık stili
            header_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")  # Yeşil
            header_font = Font(bold=True, color="FFFFFF")
            center_alignment = Alignment(horizontal="center", vertical="center")

            # Giriş tablosu - Sol tarafta (A sütunundan başla)
            ws['A1'] = f"Giriş - {self.selected_kasa_kodu} - {self.selected_ay_adi} {self.selected_yil}"
            ws['A1'].fill = header_fill
            ws['A1'].font = header_font
            ws['A1'].alignment = center_alignment
            ws.merge_cells('A1:E1')

            # Giriş tablo başlıkları
            for col_idx, header in enumerate(["Tarih", "KASA ADI", "CARI ADI", "Tutar", "Açıklama"], start=1):
                cell = ws.cell(row=2, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment

            # Giriş verileri
            for row_idx, row_data in enumerate(dataframe_to_rows(giris_df, index=False, header=False), start=3):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    # Tutar sütunu (4. sütun) için finansal format
                    if col_idx == 4 and isinstance(value, (int, float)):
                        cell.number_format = '#,##0 ₺'

            # Çıkış tablosu - Sağ tarafta (G sütunundan başla)
            header_fill_red = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")  # Kırmızı

            ws['G1'] = f"Çıkış - {self.selected_kasa_kodu} - {self.selected_ay_adi} {self.selected_yil}"
            ws['G1'].fill = header_fill_red
            ws['G1'].font = header_font
            ws['G1'].alignment = center_alignment
            ws.merge_cells('G1:K1')

            # Çıkış tablo başlıkları
            for col_idx, header in enumerate(["Tarih", "KASA ADI", "CARI ADI", "Tutar", "Açıklama"], start=7):
                cell = ws.cell(row=2, column=col_idx, value=header)
                cell.fill = header_fill_red
                cell.font = header_font
                cell.alignment = center_alignment

            # Çıkış verileri
            for row_idx, row_data in enumerate(dataframe_to_rows(cikis_df, index=False, header=False), start=3):
                for col_idx, value in enumerate(row_data, start=7):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    # Tutar sütunu (10. sütun = J) için finansal format
                    if col_idx == 10 and isinstance(value, (int, float)):
                        cell.number_format = '#,##0 ₺'

            # Sütun genişliklerini otomatik ayarla
            from openpyxl.utils import get_column_letter
            for col_idx in range(1, ws.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for row_idx in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    try:
                        if cell.value:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Excel dosyasını kaydet
            wb.save(output_path)
            self.status_label.setText(f"✅ Detaylar dışa aktarıldı: {output_path}")
            QMessageBox.information(self, "Başarılı", f"Detaylar dışa aktarıldı:\n{output_path}")

        except Exception as e:
            self.status_label.setText(f"❌ Dışa aktarma hatası: {str(e)}")
            QMessageBox.critical(self, "Hata", f"Dışa aktarma hatası:\n{str(e)}")

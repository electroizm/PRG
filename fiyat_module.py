"""
Etiket Modülü - Etiket işlemlerini yönetir
"""

import sys
import os
import re
from datetime import datetime
from pathlib import Path
import warnings
import pandas as pd
import numpy as np
import gspread
import shutil

# Üst dizini Python path'e ekle (central_config için)
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Central config import
from central_config import CentralConfigManager

from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
                             QTextEdit, QLabel, QApplication, QStatusBar, QMainWindow,
                             QTableWidget, QTableWidgetItem, QLineEdit, QCheckBox,
                             QComboBox, QMessageBox, QHeaderView, QRadioButton, QButtonGroup)
from PyQt5.QtGui import QFont, QColor
from openpyxl import load_workbook, Workbook
import pyodbc
import logging
from contextlib import contextmanager
import glob

warnings.filterwarnings('ignore')

class SapCreateThread(QThread):
    """SAP kodu oluşturma işlemlerini ayrı thread'de çalıştıran sınıf"""
    progress_update = pyqtSignal(str)
    finished_signal = pyqtSignal(bool, str)

    def __init__(self):
        super().__init__()
        # Service Account ile ayarları yükle
        self.config_manager = CentralConfigManager()

    def get_sap_data(self):
        """ID'si 8000'den büyük olan stok verilerini çeker"""
        query = """
        SELECT TOP 100 PERCENT
            sto_RECno AS [ID],
            sto_isim AS [MALZEME ADI],
            sto_kod AS [MALZEME KODU],
            dbo.fn_DepodakiMiktar(sto_kod,100,GetDate()) as DEPO,
            dbo.fn_DepodakiMiktar(sto_kod,300,GetDate()) as EXCLUSIVE,
            dbo.fn_DepodakiMiktar(sto_kod,200,GetDate()) as SUBE,
            dbo.fn_EldekiMiktar(sto_kod) AS [MIKTAR]
        FROM dbo.STOKLAR WITH (NOLOCK)
        WHERE (sto_pasif_fl IS NULL OR sto_pasif_fl=0)
            AND sto_RECno > 8000
        ORDER BY sto_kod
        """

        # PRGsheet/Ayar'dan SQL ayarlarını al
        settings = self.config_manager.get_settings()
        server = settings.get('SQL_SERVER')
        database = settings.get('SQL_DATABASE')
        username = settings.get('SQL_USERNAME')
        password = settings.get('SQL_PASSWORD')

        if not all([server, database, username, password]):
            raise ValueError("SQL bağlantı bilgileri PRGsheet/Ayar sayfasında eksik!")
        
        conn_str = f'DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password}'
        
        with pyodbc.connect(conn_str) as connection:
            cursor = connection.cursor()
            cursor.execute(query)
            
            columns = [column[0] for column in cursor.description]
            rows = cursor.fetchall()
            
            df = pd.DataFrame.from_records(rows, columns=columns)
            return df
    
    def process_sap_data(self, df):
        """SAP verilerini işler ve filtreler"""
        if df.empty:
            return df
        
        # MALZEME KODU'nun ilk 10 hanesini al
        df['MALZEME_KODU_SHORT'] = df['MALZEME KODU'].astype(str).str[:10]
        
        # Tekrar edenleri çıkar - her kod'dan sadece bir tane
        df_unique = df.drop_duplicates(subset=['MALZEME_KODU_SHORT'], keep='first')
        
        # Sadece gerekli sütunları al ve yeniden yapılandır
        result_df = pd.DataFrame()
        result_df['MALZEME KODU'] = df_unique['MALZEME_KODU_SHORT']
        result_df['MIKTAR'] = 1  # Her kod için 1 değeri
        result_df['MALZEME ADI'] = df_unique['MALZEME ADI']
        
        # MALZEME ADI'na göre küçükten büyüğe sırala
        result_df = result_df.sort_values(by='MALZEME ADI', ascending=True)
        result_df = result_df.reset_index(drop=True)
        
        return result_df
    
    def create_output_directory(self):
        """Çıktı klasörünü oluşturur"""
        output_dir = r"D:\GoogleDrive\Fiyat\SAP"
        os.makedirs(output_dir, exist_ok=True)
        return output_dir
    
    def save_split_files(self, df, output_dir):
        """Verileri 270'şer satırlık dosyalara böler ve kaydeder"""
        if df.empty:
            return
        
        chunk_size = 270
        total_rows = len(df)
        file_count = 0
        
        for i in range(0, total_rows, chunk_size):
            chunk = df.iloc[i:i+chunk_size].copy()
            file_count += 1
            filename = f"sap{file_count}.xlsx"
            filepath = os.path.join(output_dir, filename)
            
            # Excel'e kaydetmeden önce MALZEME KODU'nu sayıya çevirmeye çalış
            try:
                chunk['MALZEME KODU'] = pd.to_numeric(chunk['MALZEME KODU'])
            except (ValueError, TypeError):
                pass
            
            # ExcelWriter ile formatlama kontrolü
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                chunk.to_excel(writer, index=False, header=False, sheet_name='Sheet1')
                
                # Worksheet'i al ve MALZEME KODU sütununu "Genel" formatına ayarla
                worksheet = writer.sheets['Sheet1']
                for row in range(1, len(chunk) + 1):
                    cell = worksheet.cell(row=row, column=1)  # İlk sütun (MALZEME KODU)
                    cell.number_format = 'General'
            
            self.progress_update.emit(f"📁 {filename} dosyası oluşturuldu - {len(chunk)} satır")
        
        return file_count
        
    def run(self):
        """SAP kodlarını oluşturup Excel dosyalarına böler"""
        try:
            self.progress_update.emit("🔗 Veritabanına bağlanıyor...")
            
            # Ham verileri çek
            self.progress_update.emit("📊 ID > 8000 olan stok verileri getiriliyor...")
            df = self.get_sap_data()
            
            if df.empty:
                self.progress_update.emit("⚠️  ID > 8000 olan veri bulunamadı")
                self.finished_signal.emit(False, "ID > 8000 olan veri bulunamadı!")
                return
            
            self.progress_update.emit(f"✅ {len(df)} ham kayıt alındı, işleniyor...")
            
            # Verileri işle ve filtrele
            processed_df = self.process_sap_data(df)
            
            if processed_df.empty:
                self.progress_update.emit("⚠️  İşlenebilir veri bulunamadı")
                self.finished_signal.emit(False, "İşlenebilir veri bulunamadı!")
                return
            
            self.progress_update.emit(f"✅ {len(processed_df)} benzersiz SAP kodu hazırlandı")
            
            # Çıktı klasörünü oluştur
            output_dir = self.create_output_directory()
            self.progress_update.emit(f"📂 Çıktı klasörü hazırlandı: {output_dir}")
            
            # Dosyaları 270'şer satırlık parçalara böl ve kaydet
            self.progress_update.emit("📦 Dosyalar 270'şer satırlık parçalara bölünüyor...")
            file_count = self.save_split_files(processed_df, output_dir)
            
            self.progress_update.emit(f"🎉 İşlem tamamlandı! {file_count} dosya oluşturuldu.")
            self.finished_signal.emit(True, f"Başarılı! {len(processed_df)} kayıt {file_count} dosyaya bölündü.")
            
        except Exception as e:
            self.progress_update.emit(f"❌ Hata: {str(e)}")
            self.finished_signal.emit(False, f"SAP oluşturma hatası: {str(e)}")

class MikroFiyatThread(QThread):
    """Mikro Fiyat işlemlerini ayrı thread'de çalıştıran sınıf"""
    progress_update = pyqtSignal(str)
    finished_signal = pyqtSignal(bool, str)
    
    def __init__(self):
        super().__init__()

        self.directories = [
            r"D:\GoogleDrive\Fiyat\SAP\TOPTAN"
        ]
    
    def read_utf16_csv(self, file_path):
        """UTF-16 kodlamalı CSV dosyasını okur"""
        try:
            df = pd.read_csv(file_path, encoding='utf-16', sep='\t')
            return df
        except:
            try:
                df = pd.read_csv(file_path, encoding='utf-16le', sep='\t')
                return df
            except:
                df = pd.read_csv(file_path, encoding='utf-16', sep=';')
                return df
    
    def get_all_csv_files(self):
        """Tüm dizinlerdeki CSV dosyalarını bulur"""
        all_files = []
        for directory in self.directories:
            if os.path.exists(directory):
                pattern = os.path.join(directory, "*.csv")
                files = glob.glob(pattern)
                for file in files:
                    all_files.append({
                        'path': file,
                        'directory': directory,
                        'filename': os.path.basename(file)
                    })
            else:
                self.progress_update.emit(f"⚠️  Uyarı: {directory} dizini bulunamadı!")
        return all_files
    
    def filter_and_combine_data(self):
        """TOPTAN dizinindeki satır verisi 3 ile başlayan ve 9 karakterden uzun satırları filtreler"""
        csv_files = self.get_all_csv_files()
        
        if not csv_files:
            self.progress_update.emit("❌ Filtrelenecek CSV dosyaları bulunamadı!")
            return None
        
        filtered_toptan = []
        
        for file_info in csv_files:
            file_path = file_info['path']
            directory = file_info['directory']
            filename = file_info['filename']
            
            try:
                self.progress_update.emit(f"📄 İşleniyor: {filename}")
                df = self.read_utf16_csv(file_path)
                if df is None:
                    continue
                
                # SAP Kodu sütununu kontrol et
                sap_col = 'SAP Kodu' if 'SAP Kodu' in df.columns else 'Kalem numarası'
                if sap_col in df.columns:
                    # 3 ile başlayan ve 9 karakterden uzun olanları filtrele
                    mask = df[sap_col].astype(str).str.startswith('3') & \
                           (df[sap_col].astype(str).str.len() > 9)
                    
                    filtered_df = df[mask].copy()
                    
                    if len(filtered_df) > 0:
                        filtered_df['Kaynak_Dosya'] = filename
                        filtered_df['Tam_Yol'] = file_path
                        
                        if 'TOPTAN' in directory:
                            filtered_toptan.append(filtered_df)
                            self.progress_update.emit(f"✅ TOPTAN verisi: {len(filtered_df)} satır eklendi")
                        
            except Exception as e:
                self.progress_update.emit(f"❌ Hata - {filename}: {str(e)}")
        
        if not filtered_toptan:
            self.progress_update.emit("❌ Hiçbir satır kriterlere uymuyor!")
            return None
        
        # Excel dosyası oluştur - sadece TOPTAN sheet'i
        output_file = os.path.join(self.directories[0], 'Filtrelenmis_Veriler_3_ile_baslayan.xlsx')
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            if filtered_toptan:
                toptan_combined = pd.concat(filtered_toptan, ignore_index=True)
                toptan_combined.to_excel(writer, sheet_name='TOPTAN', index=False)
        
        # Excel dosyasını yeniden aç ve başlık satırlarını düzenle
        wb = load_workbook(output_file)
        
        if 'TOPTAN' in wb.sheetnames:
            ws = wb['TOPTAN']
            
            # İlk satırın A1 hücresini sil
            ws['A1'] = None
            
            # İlk satırdaki tüm hücreleri 1 sütun sola kaydır
            max_col = ws.max_column
            for col in range(1, max_col):
                if col < max_col:
                    ws.cell(row=1, column=col).value = ws.cell(row=1, column=col+1).value
            
            # Son sütunu temizle
            ws.cell(row=1, column=max_col).value = None
        
        wb.save(output_file)
        wb.close()
        
        return output_file
    
    def create_price_comparison(self):
        """TOPTAN verilerini işleyerek sadece gerekli sütunları içeren Excel dosyası oluşturur"""
        filter_file = os.path.join(self.directories[0], 'Filtrelenmis_Veriler_3_ile_baslayan.xlsx')
        
        if not os.path.exists(filter_file):
            self.progress_update.emit("❌ Filtrelenmis_Veriler_3_ile_baslayan.xlsx dosyası bulunamadı!")
            return None
        
        # TOPTAN verilerini oku
        try:
            toptan_df = pd.read_excel(filter_file, sheet_name='TOPTAN')
            self.progress_update.emit(f"📊 TOPTAN verileri okundu: {len(toptan_df)} satır")
        except Exception as e:
            self.progress_update.emit(f"❌ TOPTAN sheet'i okunamadı: {e}")
            return None
        
        # Sütun adlarını güncelle
        column_mapping = {
            'Ürün tanıtıcısı': 'SAP Kodu',
            'Tanım': 'Malzeme Adı', 
            'Tutar': 'TOPTAN'
        }
        
        toptan_df.rename(columns=column_mapping, inplace=True)
        
        # TOPTAN verilerinden sadece gerekli sütunları seç
        required_cols = ['SAP Kodu', 'Malzeme Adı', 'TOPTAN']
        missing_cols = [col for col in required_cols if col not in toptan_df.columns]
        if missing_cols:
            self.progress_update.emit(f"❌ TOPTAN'da eksik sütunlar: {missing_cols}")
            return None
        
        result_df = toptan_df[required_cols].copy()
        
        # Sayısal değerleri işle
        if 'TOPTAN' in result_df.columns:
            try:
                result_df['TOPTAN'] = result_df['TOPTAN'].astype(str).str.strip()
                result_df['TOPTAN'] = result_df['TOPTAN'].str.replace('.', '').str.replace(',', '.')
                result_df['TOPTAN'] = pd.to_numeric(result_df['TOPTAN'], errors='coerce')
                result_df['TOPTAN'] = result_df['TOPTAN'].round().astype('Int64')
            except Exception as e:
                self.progress_update.emit(f"❌ TOPTAN sütunu işlenirken hata: {e}")
        
        # İstatistikleri hesapla
        total_records = len(result_df)
        
        self.progress_update.emit(f"📈 Toplam kayıt: {total_records}")
        
        # Sonucu Excel'e kaydet - çalışma dizininde
        output_file = os.path.join(os.getcwd(), 'Fiyat_Mikro.xlsx')
        result_df.to_excel(output_file, index=False, engine='openpyxl')
        self.progress_update.emit(f"📁 Excel dosyası oluşturuldu: {output_file}")
        
        # Gereksiz Excel dosyalarını sil
        files_to_delete = [
            os.path.join(self.directories[0], 'tum_veriler_birlestirilmis.xlsx'),
            os.path.join(self.directories[0], 'SAP_Verileri_Ayrilmis.xlsx'),
            os.path.join(self.directories[0], 'Filtrelenmis_Veriler_3_ile_baslayan.xlsx')
        ]
        
        for file_path in files_to_delete:
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
                    self.progress_update.emit(f"🗑️  Geçici dosya silindi: {os.path.basename(file_path)}")
            except Exception as e:
                self.progress_update.emit(f"❌ Dosya silinemedi {os.path.basename(file_path)}: {e}")
        
        return result_df, output_file
    
    def upload_to_google_sheets(self, df, excel_file_path):
        """Mikro Fiyat verilerini Google Sheets'e yükler"""
        try:
            self.progress_update.emit("📊 Google Sheets yüklemesi başlatılıyor...")

            # Service Account ile Google Sheets client'ı al
            config_manager = CentralConfigManager()
            gc = config_manager.gc
            
            # Sadece gerekli sütunları seç
            required_columns = ['SAP Kodu', 'Malzeme Adı', 'TOPTAN']
            
            upload_df = df.copy()
            
            # Gerekli sütunları kontrol et
            missing_columns = [col for col in required_columns if col not in upload_df.columns]
            if missing_columns:
                self.progress_update.emit(f"❌ Eksik sütunlar: {missing_columns}")
                return False
            
            # Sadece gerekli sütunları seç
            upload_df = upload_df[required_columns].copy()
            
            # PRGsheet dosyasını aç
            spreadsheet = gc.open("PRGsheet")
            
            # Fiyat_Mikro sayfasını kontrol et ve oluştur
            try:
                worksheet = spreadsheet.worksheet('Fiyat_Mikro')
                worksheet.clear()
                self.progress_update.emit("🧹 Mevcut Fiyat_Mikro sayfası temizlendi")
            except gspread.exceptions.WorksheetNotFound:
                worksheet = spreadsheet.add_worksheet(title='Fiyat_Mikro', rows=2000, cols=10)
                self.progress_update.emit("📋 Yeni Fiyat_Mikro sayfası oluşturuldu")
            
            # Verileri yükle
            if not upload_df.empty:
                values = [upload_df.columns.values.tolist()] + upload_df.values.tolist()
                worksheet.update(values)
                self.progress_update.emit(f"✅ Google Sheets'e {len(upload_df)} satır yüklendi")
                return True
            else:
                self.progress_update.emit("⚠️  Yüklenecek veri bulunamadı")
                return False
            
        except Exception as e:
            self.progress_update.emit(f"❌ Google Sheets yükleme hatası: {e}")
            return False
    
    def run(self):
        """Mikro Fiyat ana işlem fonksiyonu"""
        try:
            self.progress_update.emit("🚀 Mikro Fiyat işlemi başlatılıyor...")
            
            # CSV dosyalarını kontrol et
            csv_files = self.get_all_csv_files()
            if not csv_files:
                self.progress_update.emit("❌ Hiçbir CSV dosyası bulunamadı!")
                self.finished_signal.emit(False, "CSV dosyası bulunamadı!")
                return
            
            self.progress_update.emit(f"📁 Toplam {len(csv_files)} CSV dosyası bulundu")
            
            # Filtreleme ve birleştirme
            filter_file = self.filter_and_combine_data()
            if not filter_file:
                self.progress_update.emit("❌ Filtreleme işlemi başarısız!")
                self.finished_signal.emit(False, "Filtreleme işlemi başarısız!")
                return
            
            self.progress_update.emit("✅ Filtreleme ve birleştirme tamamlandı")
            
            # Fiyat işleme
            result = self.create_price_comparison()
            if not result:
                self.progress_update.emit("❌ Fiyat işleme başarısız!")
                self.finished_signal.emit(False, "Fiyat işleme başarısız!")
                return
            
            result_df, excel_file = result
            
            # Google Sheets'e yükle
            upload_success = self.upload_to_google_sheets(result_df, excel_file)
            
            if upload_success:
                self.progress_update.emit("🎉 Mikro Fiyat işlemi başarıyla tamamlandı!")
                self.finished_signal.emit(True, f"Başarılı! {len(result_df)} kayıt işlendi ve Google Sheets'e yüklendi.")
            else:
                self.progress_update.emit("⚠️  Excel dosyası oluşturuldu ancak Google Sheets yüklemesi başarısız")
                self.finished_signal.emit(True, f"Kısmen başarılı! {len(result_df)} kayıt Excel'e kaydedildi.")
            
        except Exception as e:
            self.progress_update.emit(f"❌ Genel hata: {str(e)}")
            self.finished_signal.emit(False, f"Mikro Fiyat hatası: {str(e)}")

class FiyatProcessThread(QThread):
    """Fiyat işlemlerini ayrı thread'de çalıştıran sınıf"""
    progress_update = pyqtSignal(str)
    finished_signal = pyqtSignal(bool, str)
    
    def __init__(self):
        super().__init__()

    def excel_metadata_tarihini_al(self, dosya_yolu):
        """Excel dosyasının metadata'sından son kaydetme tarihini al"""
        try:
            # openpyxl ile Excel dosyasını aç
            wb = load_workbook(dosya_yolu, read_only=True)
            
            # Excel properties'den modified tarihini al
            if hasattr(wb.properties, 'modified') and wb.properties.modified:
                # Excel'in son kaydetme tarihi
                modified_date = wb.properties.modified
                wb.close()
                return modified_date.timestamp()
            
            wb.close()
            return None
            
        except Exception as e:
            return None

    def tarihe_gore_excel_dosyalarini_getir(self):
        """Excel dosylarını Excel metadata'sındaki son kaydetme tarihine göre getir (en yeni en başta)"""
        excel_dir = Path("D:/GoogleDrive/Fiyat")
        excel_files = []
        
        for file in excel_dir.glob("*.xlsx"):
            try:
                # Önce Excel metadata'sından tarihi al
                dosya_tarihi = self.excel_metadata_tarihini_al(str(file))
                
                if dosya_tarihi is None:
                    # Excel metadata'sından tarih alınamazsa, sistem dosya tarihini kullan
                    dosya_tarihi = file.stat().st_mtime
                
                excel_files.append((str(file), dosya_tarihi))
            except Exception as e:
                continue
        
        # En yeni dosyalar en başta olacak şekilde sırala (reverse=True)
        excel_files.sort(key=lambda x: x[1], reverse=True)
        return [file[0] for file in excel_files]
    
    def eski_dosyalari_sil(self):
        """7 aydan eski Excel dosyalarını Excel metadata tarihine göre sil"""
        import time
        from datetime import datetime, timedelta
        
        excel_dir = Path("D:/GoogleDrive/Fiyat")
        yedi_ay_once = datetime.now() - timedelta(days=7*30)  # 7 ay = yaklaşık 210 gün
        silinen_dosyalar = []
        
        for file in excel_dir.glob("*.xlsx"):
            try:
                # Fiyat_Listesi.xlsx dosyasını atlayalım
                if file.name == "Fiyat_Listesi.xlsx":
                    continue
                
                # Önce Excel metadata'sından tarihi al
                excel_metadata_timestamp = self.excel_metadata_tarihini_al(str(file))
                
                if excel_metadata_timestamp is not None:
                    # Excel metadata tarihi varsa onu kullan
                    dosya_tarihi = datetime.fromtimestamp(excel_metadata_timestamp)
                    tarih_kaynagi = "Excel metadata"
                else:
                    # Excel metadata tarihi yoksa sistem dosya tarihini kullan
                    mtime = file.stat().st_mtime
                    dosya_tarihi = datetime.fromtimestamp(mtime)
                    tarih_kaynagi = "Sistem dosyası"
                
                
                if dosya_tarihi < yedi_ay_once:
                    os.remove(str(file))
                    silinen_dosyalar.append(file.name)
                    self.progress_update.emit(f"🗑️  Silindi (7 aydan eski - {dosya_tarihi.strftime('%d.%m.%Y')}): {file.name}")
                
            except Exception as e:
                self.progress_update.emit(f"❌ Silinemedi: {file.name} - {e}")
        
        if silinen_dosyalar:
            self.progress_update.emit(f"📦 Toplam {len(silinen_dosyalar)} eski dosya silindi")
        else:
            self.progress_update.emit("7 aydan eski dosya bulunamadı")
        
        return silinen_dosyalar

    def sap_kodu_mu(self, deger):
        """3 ile başlayan 10 haneli SAP kodunu kontrol et"""
        if pd.isna(deger):
            return False
        
        str_deger = str(deger).strip()
        # Sadece rakamlardan oluşan, 3 ile başlayan ve 10 haneli
        if re.match(r'^3\d{9}$', str_deger):
            return True
        return False
    

    def excel_dosyasini_isle(self, dosya_yolu):
        """Excel dosyasının tüm sayfalarını işle ve SAP kodları ile fiyat bilgilerini çıkar"""
        sonuclar = []
        
        try:
            # Dosyanın tüm sayfalarını oku
            xl_dosya = pd.ExcelFile(dosya_yolu)
            
            for sayfa_adi in xl_dosya.sheet_names:
                try:
                    df = pd.read_excel(dosya_yolu, sheet_name=sayfa_adi, header=None)
                    
                    for idx, satir in df.iterrows():
                        # SAP kodu ara
                        sap_kodu = None
                        sap_adi = None
                        sap_sutun_idx = None
                        
                        for sutun_idx, hucre in enumerate(satir):
                            if self.sap_kodu_mu(hucre):
                                sap_kodu = str(hucre).strip()
                                sap_sutun_idx = sutun_idx
                                
                                # SAP adı (hemen yanındaki hücre)
                                if sutun_idx + 1 < len(satir):
                                    sonraki_hucre = satir.iloc[sutun_idx + 1]
                                    if not pd.isna(sonraki_hucre) and not self.sap_kodu_mu(sonraki_hucre):
                                        # Sayısal değer değilse isim olarak al
                                        try:
                                            float(str(sonraki_hucre))
                                        except:
                                            sap_adi = str(sonraki_hucre).strip()
                                break
                        
                        if sap_kodu:
                            # Satırdaki tüm sayısal değerleri ve konumlarını bul
                            konumlu_sayisal_degerler = []
                            
                            for sutun_idx, hucre in enumerate(satir):
                                if sutun_idx <= sap_sutun_idx:  # SAP kodu ve öncesini atla
                                    continue
                                if pd.isna(hucre):
                                    continue
                                try:
                                    if isinstance(hucre, (int, float)) and hucre > 0:
                                        konumlu_sayisal_degerler.append((sutun_idx, int(hucre)))
                                    else:
                                        str_deger = str(hucre).replace(',', '.').strip()
                                        if re.match(r'^\d*\.?\d+$', str_deger):
                                            deger = int(float(str_deger))
                                            if deger > 0:
                                                konumlu_sayisal_degerler.append((sutun_idx, deger))
                                except:
                                    continue
                            
                            if len(konumlu_sayisal_degerler) >= 2:
                                # TOPTAN: En küçük sayısal değer
                                toptan = min(konumlu_sayisal_degerler, key=lambda x: x[1])[1]
                                
                                # TOPTAN'ın pozisyonunu bul
                                toptan_konumu = None
                                for konum, deger in konumlu_sayisal_degerler:
                                    if deger == toptan:
                                        toptan_konumu = konum
                                        break
                                
                                # PERAKENDE: TOPTAN'ın hemen yanındaki sütun
                                perakende = None
                                if toptan_konumu is not None:
                                    for konum, deger in konumlu_sayisal_degerler:
                                        if konum == toptan_konumu + 1:
                                            perakende = deger
                                            break
                                
                                # Eğer perakende bulunamazsa, ikinci değeri al
                                if perakende is None and len(konumlu_sayisal_degerler) >= 2:
                                    siralanmis_degerler = sorted([v[1] for v in konumlu_sayisal_degerler])
                                    perakende = siralanmis_degerler[1]
                                
                                # LİSTE: En büyük değer
                                liste = max(konumlu_sayisal_degerler, key=lambda x: x[1])[1]
                                
                                # TOPTAN 100'den küçükse atla
                                if toptan < 100:
                                    continue
                                
                                # PERAKENDE değeri varsa kaydet
                                if perakende is not None:
                                    sonuclar.append({
                                        'SAP Kodu': sap_kodu,
                                        'Malzeme Adı': sap_adi if sap_adi else '',
                                        'TOPTAN': toptan,
                                        'PERAKENDE': perakende,
                                        'LISTE': liste,
                                        'DOSYA': os.path.basename(dosya_yolu)
                                    })
                                
                except Exception as e:
                    self.progress_update.emit(f"Sayfa işlenirken hata ({sayfa_adi}): {e}")
                    continue
                    
        except Exception as e:
            self.progress_update.emit(f"Dosya işlenirken hata ({dosya_yolu}): {e}")
        
        return sonuclar

    def fiyat_sayfasini_guncelle(self, veri):
        """PRGsheet dosyasındaki 'Fiyat' sayfasını güncelle"""
        try:
            # Service Account ile Google Sheets client'ı al
            config_manager = CentralConfigManager()
            gc = config_manager.gc

            hesap_tablosu = gc.open("PRGsheet")
            
            try:
                fiyat_sayfasi = hesap_tablosu.worksheet('Fiyat')
                fiyat_sayfasi.clear()
            except gspread.exceptions.WorksheetNotFound:
                fiyat_sayfasi = hesap_tablosu.add_worksheet(title='Fiyat', rows=1000, cols=10)
            
            if not veri.empty:
                degerler = [veri.columns.values.tolist()] + veri.values.tolist()
                fiyat_sayfasi.update(degerler)
                self.progress_update.emit(f"📊 PRGsheet 'Fiyat' sayfası güncellendi: {len(veri)} satır")
            else:
                self.progress_update.emit("Güncellenmek için veri bulunamadı")
                
        except Exception as e:
            self.progress_update.emit(f"❌ Google Sheets güncelleme hatası: {e}")
            raise

    def run(self):
        """Ana işlem fonksiyonu - Tüm fiyat işleme sürecini yönetir"""
        try:
            self.progress_update.emit("🚀 Excel dosyaları işleniyor...")
            
            # PDF dosyalarını sil
            excel_dir = Path("D:/GoogleDrive/Fiyat")
            pdf_files = list(excel_dir.glob("*.pdf"))
            if pdf_files:
                for pdf_file in pdf_files:
                    try:
                        os.remove(str(pdf_file))
                    except:
                        pass
            
            # 7 aydan eski dosyaları sil
            self.eski_dosyalari_sil()
            
            # Excel dosyalarını tarih sırasına göre al (en yeni en başta)
            tum_excel_dosyalari = self.tarihe_gore_excel_dosyalarini_getir()
            
            # Fiyat_Listesi.xlsx dosyasını listeden çıkar (çıktı dosyası olduğu için işlenmemeli)
            excel_dosyalari = [f for f in tum_excel_dosyalari if not f.endswith('Fiyat_Listesi.xlsx')]
            
            if not excel_dosyalari:
                self.progress_update.emit("❌ İşlenecek Excel dosyası bulunamadı!")
                self.finished_signal.emit(False, "Excel dosyası bulunamadı!")
                return
            
            self.progress_update.emit(f"📁 Toplam {len(excel_dosyalari)} Excel dosyası bulundu.")
            
            # SAP kodlarını ve verilerini tutacak sözlük (en güncel veri alınacak)
            sap_veri_sozlugu = {}
            veri_olan_dosyalar = []
            faydali_olmayan_dosyalar = []
            
            # Her dosyayı işle (en yeniden en eskiye - en güncel veri alınacak)
            for i, dosya_yolu in enumerate(excel_dosyalari, 1):
                self.progress_update.emit(f"⚙️  İşleniyor ({i}/{len(excel_dosyalari)}): {os.path.basename(dosya_yolu)}")
                sonuclar = self.excel_dosyasini_isle(dosya_yolu)
                
                if sonuclar:  # Dosyada veri varsa
                    dosya_adı = os.path.basename(dosya_yolu)
                    yeni_sap_sayisi = 0
                    
                    # Her SAP kodu için veriyi kaydet (sadece daha önce yoksa - en güncel önce geldiği için)
                    for sonuc in sonuclar:
                        sap_kodu = sonuc['SAP Kodu']
                        if sap_kodu not in sap_veri_sozlugu:
                            sap_veri_sozlugu[sap_kodu] = sonuc
                            yeni_sap_sayisi += 1
                    
                    if yeni_sap_sayisi > 0:
                        veri_olan_dosyalar.append(dosya_adı)
                        self.progress_update.emit(f"✅ Yeni SAP kodu eklendi: {yeni_sap_sayisi} adet - {dosya_adı}")
                    else:
                        faydali_olmayan_dosyalar.append(dosya_yolu)
                        self.progress_update.emit(f"❌ Faydalı veri yok: {dosya_adı}")
                else:
                    faydali_olmayan_dosyalar.append(dosya_yolu)
                    self.progress_update.emit(f"❌ Hiç veri yok: {os.path.basename(dosya_yolu)}")
            
            # Faydalı olmayan dosyaları sil
            if faydali_olmayan_dosyalar:
                self.progress_update.emit(f"🗑️  Faydalı olmayan {len(faydali_olmayan_dosyalar)} dosya siliniyor...")
                for dosya_yolu in faydali_olmayan_dosyalar:
                    try:
                        os.remove(dosya_yolu)
                        self.progress_update.emit(f"🗑️  Silindi (faydasız): {os.path.basename(dosya_yolu)}")
                    except Exception as e:
                        self.progress_update.emit(f"❌ Silinemedi: {os.path.basename(dosya_yolu)} - {e}")
            
            if sap_veri_sozlugu:
                # DataFrame oluştur
                tum_sonuclar = list(sap_veri_sozlugu.values())
                df = pd.DataFrame(tum_sonuclar)
                
                # Google Sheets'e kaydet
                try:
                    self.fiyat_sayfasini_guncelle(df)
                    self.progress_update.emit(f"📊 Toplam {len(df)} benzersiz SAP kodu işlendi ve PRGsheet 'Fiyat' sayfasına kaydedildi.")
                except Exception as e:
                    self.progress_update.emit(f"❌ Google Sheets'e kaydetme hatası: {e}")
                
                if veri_olan_dosyalar:
                    self.progress_update.emit(f"\n📈 Veri bulunan dosyalar ({len(veri_olan_dosyalar)} adet):")
                    for dosya_adi in veri_olan_dosyalar:
                        self.progress_update.emit(f"   ✅ {dosya_adi}")
                
                self.progress_update.emit(f"\n🎉 İşlem tamamlandı! Toplam {len(df)} SAP kodu işlendi.")
                
                self.finished_signal.emit(True, f"Başarılı! {len(df)} kayıt işlendi.")
                
            else:
                self.progress_update.emit("❌ Hiç veri bulunamadı!")
                self.finished_signal.emit(False, "Hiç veri bulunamadı!")
                
        except Exception as e:
            self.progress_update.emit(f"❌ Genel hata: {str(e)}")
            self.finished_signal.emit(False, str(e))

class FiyatModule(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Fiyat Yönetimi")
        self.setGeometry(200, 200, 800, 600)

        # Threads for processing
        self.fiyat_thread = None
        self.sap_create_thread = None
        self.mikro_fiyat_thread = None
        self.start_time = None

        # Lazy loading için flag
        self._data_loaded = False

        # PRGsheet/Ayar sayfasını yükle ve SPREADSHEET_ID'yi kontrol et
        self._check_env_file()

        self.init_ui()
        self.setup_connections()

    def showEvent(self, event):
        """Widget ilk gösterildiğinde lazy loading"""
        super().showEvent(event)
        if not self._data_loaded:
            self._data_loaded = True
            # UI render olduktan sonra - bu modülde data yok, sadece flag set et
            pass
    
    def _check_env_file(self):
        """Env dosyasını kontrol et ve SPREADSHEET_ID'yi doğrula"""
        try:
            # Service Account ile SPREADSHEET_ID'yi al
            spreadsheet_id = CentralConfigManager().MASTER_SPREADSHEET_ID
            if spreadsheet_id:
                self.spreadsheet_available = True
                return

            self.spreadsheet_available = False
        except Exception as e:
            print(f"Service Account config check error: {e}")
            self.spreadsheet_available = False
    
    def init_ui(self):
        """UI'ı başlat"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)
        
        
        # Tüm butonlar için tek bir layout
        buttons_layout = QHBoxLayout()

        # Fiyat Güncelle Button - GRİ
        self.fiyat_guncelle_button = QPushButton("Fiyat Güncelle")
        self.fiyat_guncelle_button.setFixedHeight(50)
        self.fiyat_guncelle_button.setFont(QFont("Arial Bold", 14))
        self.fiyat_guncelle_button.setStyleSheet("""
            QPushButton {
                background-color: #808080;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 8px;
                font-weight: bold;
                font-size: 16px;
            }
            QPushButton:hover {
                background-color: #999999;
            }
            QPushButton:pressed {
                background-color: #666666;
            }
            QPushButton:disabled {
                background-color: #444444;
            }
        """)

        # SAP Kodu Oluştur Button - GRİ
        self.sap_create_button = QPushButton("SAP Kodu Oluştur")
        self.sap_create_button.setFixedHeight(50)
        self.sap_create_button.setFont(QFont("Arial Bold", 14))
        self.sap_create_button.setStyleSheet("""
            QPushButton {
                background-color: #808080;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 8px;
                font-weight: bold;
                font-size: 16px;
            }
            QPushButton:hover {
                background-color: #999999;
            }
            QPushButton:pressed {
                background-color: #666666;
            }
            QPushButton:disabled {
                background-color: #444444;
            }
        """)

        # Mikro Fiyat Oluştur Button - GRİ
        self.mikro_fiyat_button = QPushButton("Mikro Fiyat Oluştur")
        self.mikro_fiyat_button.setFixedHeight(50)
        self.mikro_fiyat_button.setFont(QFont("Arial Bold", 14))
        self.mikro_fiyat_button.setStyleSheet("""
            QPushButton {
                background-color: #808080;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 8px;
                font-weight: bold;
                font-size: 16px;
            }
            QPushButton:hover {
                background-color: #999999;
            }
            QPushButton:pressed {
                background-color: #666666;
            }
            QPushButton:disabled {
                background-color: #444444;
            }
        """)

        # Butonları layout'a ekle
        buttons_layout.addWidget(self.fiyat_guncelle_button)
        buttons_layout.addWidget(self.sap_create_button)
        buttons_layout.addWidget(self.mikro_fiyat_button)

        layout.addLayout(buttons_layout)

        self.console_output = QTextEdit()
        self.console_output.setReadOnly(True)
        self.console_output.setFont(QFont("Consolas", 12))
        self.console_output.setStyleSheet("""
            QTextEdit {
                background-color: #1a1a1a;
                color: #00ff00;
                border: 2px solid #404040;
                border-radius: 8px;
                padding: 10px;
                font-family: 'Consolas', 'Courier New', monospace;
            }
        """)
        self.console_output.setPlaceholderText("Konsol çıktıları burada görünecek...")

        # Progress Bar ve Status Label - Kompakt ve yan yana
        from PyQt5.QtWidgets import QProgressBar
        status_layout = QHBoxLayout()

        self.status_label = QLabel("Hazır")
        self.status_label.setStyleSheet("""
            QLabel {
                color: #cccccc;
                padding: 4px 8px;
                background-color: #2d2d2d;
                border-top: 1px solid #404040;
                font-size: 11px;
                max-height: 20px;
            }
        """)

        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 1px solid #2d2d2d;
                border-radius: 3px;
                background-color: #1a1a1a;
                color: #ffffff;
                text-align: center;
                font-weight: bold;
                min-height: 25px;
                max-height: 25px;
                font-size: 17px;
            }
            QProgressBar::chunk {
                background-color: #4CAF50;
                border-radius: 3px;
            }
        """)

        status_layout.addWidget(self.status_label, 1)
        status_layout.addWidget(self.progress_bar, 6)  # 3 katına çıkarıldı (2'den 6'ya)

        status_widget = QWidget()
        status_widget.setLayout(status_layout)
        status_widget.setStyleSheet("background-color: #2d2d2d; border-top: 1px solid #404040;")

        # Layout arrangement - Zaten yukarıda eklendi, tekrar eklemeyelim
        layout.addWidget(self.console_output, 1)  # Expandable
        layout.addWidget(status_widget)
        
        # Widget'ın genel stilini ayarla
        self.setStyleSheet("""
            QWidget {
                background-color: #1a1a1a;
                color: #ffffff;
            }
        """)
    
    def setup_connections(self):
        """Bağlantıları kur"""
        self.fiyat_guncelle_button.clicked.connect(self.fiyat_guncelle)
        self.sap_create_button.clicked.connect(self.sap_create)
        self.mikro_fiyat_button.clicked.connect(self.mikro_fiyat_create)
    
    def fiyat_guncelle(self):
        """Fiyat güncelleme işlemini başlat"""
        # Fiyat dizini var mı kontrol et
        fiyat_dir = Path("D:/GoogleDrive/Fiyat")
        if not fiyat_dir.exists():
            self.print_to_console(f"HATA: Fiyat dizini bulunamadı: {fiyat_dir}")
            self.status_label.setText("Hata: Fiyat dizini bulunamadı")
            return
        
        # Environment değişkenlerini kontrol et
        if not self.spreadsheet_available:
            self.print_to_console("HATA: PRGsheet/Ayar sayfasında gerekli bilgiler bulunamadı")
            self.status_label.setText("Hata: Environment değişkenleri eksik")
            return
        
        self.print_to_console("Fiyat güncelleme işlemi başlatılıyor...")
        self.status_label.setText("Fiyat işlemi çalışıyor...")
        self.set_buttons_enabled(False)
        
        # Başlangıç zamanını kaydet
        self.start_time = datetime.now()
        
        # Thread'i başlat
        self.fiyat_thread = FiyatProcessThread()
        self.fiyat_thread.progress_update.connect(self.on_progress_update)
        self.fiyat_thread.finished_signal.connect(self.on_process_finished)
        self.fiyat_thread.start()
    
    def on_progress_update(self, message):
        """Thread'den gelen progress mesajlarını işle"""
        self.print_to_console(message)
    
    def on_process_finished(self, success, message):
        """Thread tamamlandığında çağrılan method"""
        # Çalışma süresini hesapla
        if self.start_time:
            end_time = datetime.now()
            duration = end_time - self.start_time
            duration_seconds = duration.total_seconds()
            
            # Süreyi formatla
            if duration_seconds < 60:
                duration_str = f"{duration_seconds:.1f} saniye"
            else:
                minutes = int(duration_seconds // 60)
                seconds = duration_seconds % 60
                duration_str = f"{minutes} dakika {seconds:.1f} saniye"
            
            if success:
                self.print_to_console(f"✅ İşlem başarıyla tamamlandı! Süre: {duration_str}")
                self.status_label.setText(f"İşlem tamamlandı - Süre: {duration_str}")
            else:
                self.print_to_console(f"❌ İşlem başarısız oldu: {message} - Süre: {duration_str}")
                self.status_label.setText(f"İşlem başarısız - Süre: {duration_str}")
        else:
            if success:
                self.print_to_console("✅ İşlem başarıyla tamamlandı!")
                self.status_label.setText("İşlem tamamlandı")
            else:
                self.print_to_console(f"❌ İşlem başarısız oldu: {message}")
                self.status_label.setText("İşlem başarısız")
        
        self.set_buttons_enabled(True)
        self.fiyat_thread = None
    
    
    def print_to_console(self, message):
        """Konsola mesaj yazdır"""
        self.console_output.append(f"[{self.get_timestamp()}] {message}")
        # Otomatik olarak en alta kaydır
        cursor = self.console_output.textCursor()
        cursor.movePosition(cursor.End)
        self.console_output.setTextCursor(cursor)
        # Anlık görüntüleme için UI'ı zorla güncelle
        QApplication.processEvents()
    
    def get_timestamp(self):
        """Zaman damgası al"""
        return datetime.now().strftime("%H:%M:%S")
    
    def sap_create(self):
        """SAP kodları oluşturma işlemini başlat"""
        # Environment değişkenlerini kontrol et

        required_vars = ['SQL_SERVER', 'SQL_DATABASE', 'SQL_USERNAME', 'SQL_PASSWORD']
        missing_vars = [var for var in required_vars if not os.getenv(var)]
        
        if missing_vars:
            self.print_to_console(f"HATA: PRGsheet/Ayar sayfasında eksik değişkenler: {', '.join(missing_vars)}")
            self.status_label.setText("Hata: Environment değişkenleri eksik")
            return
        
        self.print_to_console("SAP kodları oluşturma işlemi başlatılıyor...")
        self.status_label.setText("SAP oluşturma işlemi çalışıyor...")
        self.set_buttons_enabled(False)
        
        # Başlangıç zamanını kaydet
        self.start_time = datetime.now()
        
        # Thread'i başlat
        self.sap_create_thread = SapCreateThread()
        self.sap_create_thread.progress_update.connect(self.on_progress_update)
        self.sap_create_thread.finished_signal.connect(self.on_sap_create_finished)
        self.sap_create_thread.start()
    
    def on_sap_create_finished(self, success, message):
        """SAP create thread tamamlandığında çağrılan method"""
        # Çalışma süresini hesapla
        if self.start_time:
            end_time = datetime.now()
            duration = end_time - self.start_time
            duration_seconds = duration.total_seconds()
            
            # Süreyi formatla
            if duration_seconds < 60:
                duration_str = f"{duration_seconds:.1f} saniye"
            else:
                minutes = int(duration_seconds // 60)
                seconds = duration_seconds % 60
                duration_str = f"{minutes} dakika {seconds:.1f} saniye"
            
            if success:
                self.print_to_console(f"✅ SAP oluşturma işlemi başarıyla tamamlandı! Süre: {duration_str}")
                self.status_label.setText(f"SAP oluşturma tamamlandı - Süre: {duration_str}")
            else:
                self.print_to_console(f"❌ SAP oluşturma işlemi başarısız oldu: {message} - Süre: {duration_str}")
                self.status_label.setText(f"SAP oluşturma başarısız - Süre: {duration_str}")
        else:
            if success:
                self.print_to_console("✅ SAP oluşturma işlemi başarıyla tamamlandı!")
                self.status_label.setText("SAP oluşturma tamamlandı")
            else:
                self.print_to_console(f"❌ SAP oluşturma işlemi başarısız oldu: {message}")
                self.status_label.setText("SAP oluşturma başarısız")
        
        self.set_buttons_enabled(True)
        self.sap_create_thread = None
    
    def mikro_fiyat_create(self):
        """Mikro Fiyat oluşturma işlemini başlat"""
        # CSV dizinini kontrol et
        directories = [
            r"D:\GoogleDrive\Fiyat\SAP\TOPTAN"
        ]
        
        missing_dirs = [d for d in directories if not os.path.exists(d)]
        if missing_dirs:
            self.print_to_console(f"HATA: Gerekli dizinler bulunamadı: {', '.join(missing_dirs)}")
            self.status_label.setText("Hata: CSV dizinleri eksik")
            return
        
        # Environment değişkenlerini kontrol et (Google Sheets için)

        credentials_file = os.getenv('GOOGLE_SHEETS_CREDENTIALS_FILE')
        if not credentials_file:
            self.print_to_console("UYARI: Google Sheets kimlik bilgileri bulunamadı. Sadece Excel dosyası oluşturulacak.")
        
        self.print_to_console("Mikro Fiyat oluşturma işlemi başlatılıyor...")
        self.status_label.setText("Mikro Fiyat işlemi çalışıyor...")
        self.set_buttons_enabled(False)
        
        # Başlangıç zamanını kaydet
        self.start_time = datetime.now()
        
        # Thread'i başlat
        self.mikro_fiyat_thread = MikroFiyatThread()
        self.mikro_fiyat_thread.progress_update.connect(self.on_progress_update)
        self.mikro_fiyat_thread.finished_signal.connect(self.on_mikro_fiyat_finished)
        self.mikro_fiyat_thread.start()
    
    def on_mikro_fiyat_finished(self, success, message):
        """Mikro Fiyat thread tamamlandığında çağrılan method"""
        # Çalışma süresini hesapla
        if self.start_time:
            end_time = datetime.now()
            duration = end_time - self.start_time
            duration_seconds = duration.total_seconds()
            
            # Süreyi formatla
            if duration_seconds < 60:
                duration_str = f"{duration_seconds:.1f} saniye"
            else:
                minutes = int(duration_seconds // 60)
                seconds = duration_seconds % 60
                duration_str = f"{minutes} dakika {seconds:.1f} saniye"
            
            if success:
                self.print_to_console(f"✅ Mikro Fiyat işlemi başarıyla tamamlandı! Süre: {duration_str}")
                self.status_label.setText(f"Mikro Fiyat tamamlandı - Süre: {duration_str}")
            else:
                self.print_to_console(f"❌ Mikro Fiyat işlemi başarısız oldu: {message} - Süre: {duration_str}")
                self.status_label.setText(f"Mikro Fiyat başarısız - Süre: {duration_str}")
        else:
            if success:
                self.print_to_console("✅ Mikro Fiyat işlemi başarıyla tamamlandı!")
                self.status_label.setText("Mikro Fiyat tamamlandı")
            else:
                self.print_to_console(f"❌ Mikro Fiyat işlemi başarısız oldu: {message}")
                self.status_label.setText("Mikro Fiyat başarısız")
        
        self.set_buttons_enabled(True)
        self.mikro_fiyat_thread = None

    def set_buttons_enabled(self, enabled):
        """Butonları aktif/pasif yap"""
        self.fiyat_guncelle_button.setEnabled(enabled)
        self.sap_create_button.setEnabled(enabled)
        self.mikro_fiyat_button.setEnabled(enabled)

    def closeEvent(self, event):
        """Pencere kapatılırken thread'leri temizle"""
        if self.fiyat_thread and self.fiyat_thread.isRunning():
            self.fiyat_thread.terminate()
            self.fiyat_thread.wait()
        if self.sap_create_thread and self.sap_create_thread.isRunning():
            self.sap_create_thread.terminate()
            self.sap_create_thread.wait()
        if self.mikro_fiyat_thread and self.mikro_fiyat_thread.isRunning():
            self.mikro_fiyat_thread.terminate()
            self.mikro_fiyat_thread.wait()
        event.accept()